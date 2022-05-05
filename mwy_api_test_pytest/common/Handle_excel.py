"""
=============================
Author:sangshen
time:2022/2/12 
E-mail:928857196@qq.com 
=============================
"""
'''
封装的需求：
    1、数据读取：封装一个可以读取任意excel文件的方法，可以选指定读取的表单
    2、数据写入：
        文件名：
        表单：
        行：
        列：
        写入的值：
'''

import openpyxl


class HandleExcel:

    def __init__(self, filename, sheet_name):
        '''

        :param filename: 文件路径
        :param sheet_name: 表单名
        '''
        self.filename = filename
        self.sheet_name = sheet_name

    # def data_read(self):
    #     """读取文件"""
    #     workbook = openpyxl.load_workbook(filename)
    #     sh = workbook[sheet_name]
    #     res = list(sh.rows)
    #     title = [i.value for i in res[0]]
    #     cases = []
    #     for item in res[1:]:
    #         ress = [i.value for i in item]
    #         dic = dict(zip(title, ress))
    #         cases.append(dic)
    #     return cases

    def data_read(self):
        '''
        :return:
        '''
        """读取excel文件中的数据"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheet_name]
        # 获取第一行内容
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行以外的其他行
        for item in res[1:]:
            ress = [i.value for i in item]
            dic = dict(zip(title, ress))
            cases.append(dic)
        # 返回读取出来的数据
        return cases

    def data_write(self, row, column, value):
        '''
        :param row: 写入的行
        :param column: 写入的列
        :param value: 写入的值
        :return:
        '''
        """往文件中写入数据"""
        # 打开excel文件
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheet_name]

        # 往excel中写入数据：指定要写入数据的表格所在的行和列
        sh.cell(row=row, column=column, value=value)

        # 把当前文件中的内容重新
        workbook.save(self.filename)
